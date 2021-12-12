import bs4
import requests
from openpyxl import Workbook

dic_pegatina = {
    'imagen': "",
    'nombre': "",
    'autor': "",
    'precio': "",
}


def cargar_pagina(number):
    html = requests.get("https://www.redbubble.com/es/shop/stickers?page=" + str(number)).content
    soup = bs4.BeautifulSoup(html, "html.parser")
    return soup


def cargar_elementos():
    list_pegatinas = list()
    for i in range(10):
        soup = cargar_pagina(i)
        list_elementos = soup.findAll("a", {"class": "styles__link--3QJ5N"})
        for elemento in list_elementos:
            url = elemento.find("div", {"class": "styles__box--2Ufmy styles__ratioInner--miSrD"}).find("img").attrs[
                "src"]
            nombre = elemento.find("span", {
                "class": "styles__box--2Ufmy styles__text--23E5U styles__display6--3wsBG styles__nowrap--33UtL styles__display-block--3kWC4"}).text
            autor = list(elemento.find("div", {
                "class": "styles__box--2Ufmy styles__disableLineHeight--6s16u styles__marginTop-xxs--Ncwjw"}).find(
                "span", {
                    "class": "styles__box--2Ufmy styles__text--23E5U styles__body2--2dvwJ styles__nowrap--33UtL styles__display-block--3kWC4"}))[
                0]
            precio = elemento.find("span", {
                "class": "styles__box--2Ufmy styles__text--23E5U styles__display5--2KoKo styles__display-block--3kWC4"}).find(
                "span").find("span").text
            precio = precio[:-1]
            precio = float(precio.replace(",", "."))

            nueva_pegatina = dic_pegatina.copy()
            nueva_pegatina["imagen"] = url
            nueva_pegatina["nombre"] = nombre
            nueva_pegatina["autor"] = autor
            nueva_pegatina["precio"] = precio
            list_pegatinas.append(nueva_pegatina)

    return (list_pegatinas)




def pegatinas_excel(name):
    list_pegatinas = cargar_elementos()
    doc = Workbook()
    hoja_nueva = doc.create_sheet("hoja")

    for index, key in enumerate(dic_pegatina.keys()):
        hoja_nueva.cell(row=1, column=index + 1, value=key)

    for fila, nueva_pegatina in enumerate(list_pegatinas):
        for columna, valor in enumerate(nueva_pegatina.values()):
            hoja_nueva.cell(row=fila + 2, column=columna + 1, value= valor)
    doc.save(name)





